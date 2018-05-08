#!/usr/bin/env python

# You'll need to pip import openpyxl
import openpyxl

excelFile = openpyxl.load_workbook('stefan-data.xlsx')
sheet1 = excelFile.active

convert = {
    '0': '<10',
    '0.5': 10,
    '1': 20,
    '1.5': 30,
    '2': 40,
    '2.5': 60,
    '3': 80,
    '3.5': 120,
    '4': 160,
    '4.5': 240,
    '5': 320,
    '5.5': 480,
    '6': 640,
    '6.5': 960,
    '7': 1280,
    '7.5': 1920,
    '8.5': 2560,
}

# Option 1
for row in sheet1.iter_rows():
    for cell in row:
        value = str(cell.value)
        if value in convert:
            cell.value = convert[value]

# Option 2
for row in sheet1.iter_rows():
    for cell in row:
        cell.value = convert.get(str(cell.value), cell.value)

# Option 3
for row in sheet1.iter_rows():
    for cell in row:
        try:
            cell.value = convert[str(cell.value)]
        except KeyError:
            pass

excelFile.save('stefan-new-data.xlsx')
