
### Script to edit excel spreadsheet using openpyxl
'''This script extracts information from cells in a spreadsheet and writes them to empty columns. It also re-arranges the 
positions of various cells which makes the spreadsheet more suitable for downstream analysis. 

The spreadsheet contained raw pixel values of UV and non-UV images of nestling birds.
'''

import openpyxl

# Load excel file
wb = openpyxl.load_workbook('Raw pixel values for mouth markings.xlsx')

# Specify sheet of interest
sheet = wb.get_sheet_by_name('Sheet1')

### Some rows in original excel table were empty. These needed to be removed.
# Remove empty rows
for i in range (2, sheet.max_row + 1):
    if sheet.cell (row = i, column = 1).value == None:
        sheet.delete_rows(i)

### All metadata (individual ID, photo ID, UV or visible image) for each entry was contained in a single cell in the "Individual" column.
## I wanted to separate this out so that each bit of information was in a separate column. I created empty columns in excel and then filled them out in openpyxl
# Separate out information contained in "Individual" column (column 2) into separate columns
for i in range (2, sheet.max_row + 1):
    # Create cell object containing all the information
    cell = sheet.cell (row = i, column = 2).value
    
    # Obtain information for individual ID, Photo ID and whether visual or UV image
    individual = "_".join (cell.split ('_')[:-4]) # Split cell on '_' and return everything but last 4 elements to get individual identifier
    photo_ID = "_".join (cell.split ('_')[-4:-2])
    UV_or_vis =  cell.split ('_')[-2]
    
    # Assign each bit of information to correct column
    sheet.cell (row = i, column = 2).value = individual # Replace old cell entry with new individual value
    sheet.cell (row = i, column = 3).value = photo_ID
    sheet.cell (row = i, column = 4).value = UV_or_vis

### To convert raw pixel values to avian colour vision you need UV and non-UV values for a single individual all on the same row.
## Initially UV values and non-UV values were on different rows

# Move UV raw pixel values onto same line as Vis raw pixel values of corresponding individual
for i in range (2, sheet.max_row + 1):
    if sheet.cell (row = i, column = 4).value == "UV":
        # Collect UV image parameters
        UV_lumMean = sheet.cell (row = i, column = 6).value
        UV_lumSD = sheet.cell (row = i, column = 7).value
        UV_area = sheet.cell (row = i, column = 8).value
        UV_RedLinearMean = sheet.cell (row = i, column = 9).value
        UV_RedLinearSD = sheet.cell (row = i, column = 10).value
        UV_GreenLinearMean = sheet.cell (row = i, column = 11).value
        UV_GreenLinearSD = sheet.cell (row = i, column = 12).value
        UV_BlueLinearMean = sheet.cell (row = i, column = 13).value
        UV_BlueLinearSD = sheet.cell (row = i, column = 14).value
        
        # Write UV image parameters to correct UV columns
        sheet.cell (row = i-8, column = 15).value = UV_lumMean
        sheet.cell (row = i-8, column = 16).value = UV_lumSD
        sheet.cell (row = i-8, column = 17).value = UV_area
        sheet.cell (row = i-8, column = 18).value = UV_RedLinearMean
        sheet.cell (row = i-8, column = 19).value = UV_RedLinearSD
        sheet.cell (row = i-8, column = 20).value = UV_GreenLinearMean
        sheet.cell (row = i-8, column = 21).value = UV_GreenLinearSD
        sheet.cell (row = i-8, column = 22).value = UV_BlueLinearMean
        sheet.cell (row = i-8, column = 23).value = UV_BlueLinearSD

### Having copied UV values onto same rows as Vis values, I needed to delete the rows that contained UV values.
## You can't do this with a for loop because, after deleting a row, the for loop index is now out-of-date and, when you have multiple UV entries after one another the loop misses out every second one.
## Instead, a while loop allows you to have more control over where you are in the loop by either going one step forward or backwards depending on whether a certain condition is met. 
## This approach was not needed earlier on in the script when deleting empty rows as they only occurred 8 lines apart from each other (due to structure of database)

n=2
while n <= sheet.max_row:
    if sheet.cell (row = n, column = 4).value == "UV":
        sheet.delete_rows(n)
        n = n - 1
    else:
        n += 1
    
### Now that UV and Vis are on same line, there is no need to have a column specifying UV or Vis (as both bits of information are on the same row)    
# Remove UVorVis column
sheet.delete_cols (4)

### Check how many individuals in dataset
## There are multiple entries for each individual but focussing on different regions of interest
names = []
for i in range (2, sheet.max_row + 1):
    individual = sheet.cell (row = i, column = 2).value
    names.append (individual)
len (set(names)) # Returns only unique elements of the "names" list  


# Save changes
wb.save('Raw pixels neatened.xlsx')

